"""Build Use Case 1's assets: one messy claims CSV, the VBA's expected clean
output (a Python oracle that mirrors the macro rule-for-rule), and the Excel
workbook shell.

The story: an actuary receives a monthly claims listing (bordereau) from a
third-party administrator as a CSV, runs an old Excel macro that cleans and
enriches it, and exports the result for downstream use.

Why ~200k rows: it opens in Excel, but the row-by-row VBA macro takes a
couple of minutes to grind through it — that's the point. The Databricks
notebook does the identical work in seconds.

The CSV is deliberately messy, the way real vendor extracts are:
  - loss dates in three formats (dd/mm/yyyy, yyyy-mm-dd, dd-Mon-yy) plus a
    few unusable ones (TBC, blank, impossible dates)
  - money as "£1234.56", "1234.56", "-", "" and "(123.45)" for negatives
  - status codes with dirt: O / RO / C / CWP plus "OPEN", "o", "C ", "closed"
  - ~2,000 exact duplicate rows (the vendor extract double-fires)

The VBA (and this oracle) do the same clean + enrich:
  keep first row per ClaimRef → drop rows whose loss date won't parse
  (silently!) → strip £/() → map status codes → drop the Handler column →
  incurred = round(paid + outstanding, 2) → large_loss_flag = incurred > 100k

Run:  uv run --with pandas --with openpyxl python build_demo_assets.py
Outputs are committed so the repo is deterministic and self-contained.
"""
import csv
import os
import random
from datetime import date, timedelta

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data"))
os.makedirs(DATA, exist_ok=True)

MONTHS_3 = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_NO = {m.upper(): i + 1 for i, m in enumerate(MONTHS_3)}

PERILS = ["AD", "TPPD", "TPBI", "Theft", "Windscreen", "Fire", "Flood", "Storm"]
PERIL_W = [0.30, 0.22, 0.10, 0.08, 0.12, 0.06, 0.06, 0.06]
STATUS_POOL = [("C", 0.42), ("O", 0.24), ("CWP", 0.09), ("RO", 0.05),
               ("closed", 0.07), ("OPEN", 0.06), ("o", 0.03), ("C ", 0.04)]
BAD_LOSS_DATES = ["TBC", "", "31/13/2025", "00/01/2025", "31/04/2025", "unknown"]

N_BASE = 198_000          # unique claims
N_DUPES = 2_000           # exact duplicate rows appended adjacent
BAD_DATE_RATE = 0.004     # ~800 rows the VBA silently drops
LARGE_LOSS = 100_000.0    # incurred above this is flagged a large loss


def fmt_date(d, style):
    if style == 0:
        return d.strftime("%d/%m/%Y")
    if style == 1:
        return d.strftime("%Y-%m-%d")
    return f"{d.day:02d}-{MONTHS_3[d.month - 1]}-{d.year % 100:02d}"


def fmt_amount(value, rng):
    if value < 0:
        return f"({abs(value):.2f})"
    if value == 0:
        return rng.choice(["-", "", "0.00"])
    return f"£{value:.2f}" if rng.random() < 0.6 else f"{value:.2f}"


def gen_rows(seed):
    rng = random.Random(seed)
    month_end = date(2025, 11, 30)
    rows = []
    for i in range(1, N_BASE + 1):
        claim_ref = f"CLM-2025-{i:07d}"
        policy_ref = f"POL-{rng.randint(1_000_000, 9_999_999)}"
        back = int(rng.betavariate(1.2, 3.0) * 900)
        loss = month_end - timedelta(days=back)
        report = loss + timedelta(days=int(rng.expovariate(1 / 14)))
        if report > month_end:
            report = month_end

        if rng.random() < BAD_DATE_RATE:
            loss_s = rng.choice(BAD_LOSS_DATES)
        else:
            loss_s = fmt_date(loss, rng.choices([0, 1, 2], weights=[0.7, 0.2, 0.1])[0])
        report_s = "" if rng.random() < 0.003 else report.strftime("%d/%m/%Y")

        status = rng.choices([s for s, _ in STATUS_POOL],
                             weights=[w for _, w in STATUS_POOL])[0]
        clean_status = status.strip().upper()

        # severity: mostly small, occasional large loss
        if rng.random() < 0.006:
            sev = round(rng.lognormvariate(11.9, 0.5), 2)   # large losses ~ £150k+
        else:
            sev = round(rng.lognormvariate(7.6, 0.9), 2)    # ~£2k median
        if clean_status == "CWP":            # closed without payment
            paid_v, outst_v = 0.0, 0.0
        elif clean_status in ("C", "CLOSED"):
            paid_v, outst_v = sev, 0.0
        else:                                 # open / reopened
            frac = rng.uniform(0.0, 0.7)
            paid_v = round(sev * frac, 2)
            outst_v = round(sev - paid_v, 2)
        if rng.random() < 0.01:               # salvage/subrogation recovery
            paid_v = round(-abs(sev) * rng.uniform(0.05, 0.3), 2)
            outst_v = 0.0

        rows.append([
            claim_ref, policy_ref, loss_s, report_s, status,
            rng.choices(PERILS, weights=PERIL_W)[0],
            "".join(rng.choices("ABCDEFGHJKLMNPRSTW", k=3)),
            fmt_amount(paid_v, rng), fmt_amount(outst_v, rng),
        ])

    for idx in sorted(rng.sample(range(len(rows)), N_DUPES), reverse=True):
        rows.insert(idx + 1, list(rows[idx]))
    return rows


# ---------------------------------------------------------------------------
# The oracle — EXACTLY the VBA's rules (see ClaimsBordereauETL.bas)
# ---------------------------------------------------------------------------
def parse_date_iso(s):
    s = s.strip()
    if not s:
        return ""
    try:
        if "/" in s:
            p = s.split("/")
            if len(p) != 3:
                return ""
            d, m, y = int(p[0]), int(p[1]), int(p[2])
        elif "-" in s:
            p = s.split("-")
            if len(p) != 3:
                return ""
            if len(p[0]) == 4:
                y, m, d = int(p[0]), int(p[1]), int(p[2])
            else:
                d = int(p[0]); m = MONTH_NO.get(p[1][:3].upper(), 0); y = 2000 + int(p[2])
                if m == 0:
                    return ""
        else:
            return ""
        if not (1 <= m <= 12 and 1 <= d <= 31):
            return ""
        return date(y, m, d).isoformat()
    except (ValueError, TypeError):
        return ""


def parse_amount(s):
    t = s.strip()
    if t in ("", "-"):
        return 0.0
    neg = t.startswith("(") and t.endswith(")")
    if neg:
        t = t[1:-1]
    t = t.replace("£", "").replace(",", "")
    try:
        v = float(t)
    except ValueError:
        v = 0.0
    return -v if neg else v


def map_status(s):
    u = s.strip().upper()
    return {"O": "Open", "OPEN": "Open", "RO": "Reopened", "REOPENED": "Reopened",
            "C": "Closed", "CLOSED": "Closed", "CWP": "ClosedWithoutPayment"}.get(u, "UNKNOWN")


def run_oracle(rows):
    seen, out = set(), []
    for r in rows:
        ref = r[0].strip()
        if ref in seen:
            continue
        seen.add(ref)
        loss = parse_date_iso(r[2])
        if not loss:
            continue                       # the silent drop
        paid = parse_amount(r[7])
        outst = parse_amount(r[8])
        incurred = round(paid + outst, 2)
        out.append([ref, r[1].strip(), loss, parse_date_iso(r[3]),
                    map_status(r[4]), r[5].strip(),
                    f"{paid:.2f}", f"{outst:.2f}", f"{incurred:.2f}",
                    "Y" if incurred > LARGE_LOSS else "N"])
    return out


IN_HEADER = ["ClaimRef", "PolicyRef", "LossDate", "ReportDate", "Status",
             "Peril", "Handler", "PaidGBP", "OutstandingGBP"]
OUT_HEADER = ["claim_ref", "policy_ref", "loss_date", "report_date", "status",
              "peril", "paid_gbp", "outstanding_gbp", "incurred_gbp", "large_loss_flag"]


def write_csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    size = os.path.getsize(path) / 1e6
    print(f"✓ {os.path.basename(path):32s} {len(rows):>8,} rows  {size:5.1f} MB")


rows = gen_rows(seed=71)
write_csv(os.path.join(DATA, "claims_raw.csv"), IN_HEADER, rows)
oracle = run_oracle(rows)
write_csv(os.path.join(DATA, "claims_clean_excel_output.csv"), OUT_HEADER, oracle)
dropped = len({r[0] for r in rows}) - len(oracle)
large = sum(1 for r in oracle if r[9] == "Y")
print(f"   {len(rows) - len({tuple(r) for r in rows})} duplicate rows removed, "
      f"{dropped} rows silently dropped (bad loss dates), {large} large losses flagged")

# ---------------------------------------------------------------------------
# Workbook shell — import the .bas and save as .xlsm
# ---------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = "Instructions"
ws["A1"] = "Claims Bordereau ETL — legacy workbook (demo fixture)"
ws["A1"].font = Font(bold=True, size=14)
for i, line in enumerate([
    "",
    "One-time setup:",
    "  1. Open this file in Excel.",
    "  2. Developer → Visual Basic (or Alt+F11 / Option+F11).",
    "  3. File → Import File… → select ClaimsBordereauETL.bas (next to this file).",
    "  4. Save As → Excel Macro-Enabled Workbook → ClaimsBordereauETL.xlsm.",
    "",
    "To run the monthly clean (what the actuary does):",
    "  1. Run macro CleanBordereau (Developer → Macros → Run).",
    "  2. Pick the month's CSV (data/claims_raw.csv).",
    "  3. Watch it grind row by row for a couple of minutes — that's the point.",
    "  4. It writes <input>_CLEAN.csv next to the input; that's the file that",
    "     normally gets uploaded into the pricing / reserving system.",
    "",
    "About this demo: all data is synthetic; no customer data is used.",
], start=2):
    ws.cell(row=i, column=1, value=line)
ws.column_dimensions["A"].width = 92
out = os.path.join(HERE, "ClaimsBordereauETL.xlsx")
wb.save(out)
print(f"✓ {os.path.basename(out)}")
