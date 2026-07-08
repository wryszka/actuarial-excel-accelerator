"""Build Use Case 2's source assets.

Produces:
  ../data/sf_inputs.csv           ~100 entities' balance-sheet inputs (ENT-001 = the workbook entity)
  ../data/calibration_2025.json   the 2025 parameter set (shocks, sigmas, correlations)
  ../data/calibration_2026.json   the 2026 update (several parameters move)
  ../data/expected_entity_001.json  oracle: ENT-001's SCR breakdown under both calibrations
  SF_Model.xlsx                   the "before" — a one-entity Standard Formula
                                  workbook with live formulas (Inputs, Calibration,
                                  Model, Output tabs)

The model is a deliberately simple three-module Standard Formula shape —
non-life premium & reserve risk, market interest-rate risk, catastrophe —
aggregated with a correlation matrix, plus an operational-risk add-on.
Simple enough to read on one screen; real enough to be recognisable.

    SCR_nl  = 3 · sqrt((σp·Vp)² + 2·ρpr·(σp·Vp)·(σr·Vr) + (σr·Vr)²)
    SCR_mkt = |assets·dur_a − BEL·dur_l| · ir_shock
    SCR_cat = cat_factor · Vp
    BSCR    = sqrt(Σᵢⱼ ρᵢⱼ·SCRᵢ·SCRⱼ)
    Op      = min(op_factor · Vp, 0.30 · BSCR)
    SCR     = BSCR + Op

Run:  uv run --with openpyxl python build_uc2_assets.py
Outputs are committed so the repo is deterministic and self-contained.
"""
import csv
import json
import math
import os
import random

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data"))
os.makedirs(DATA, exist_ok=True)

CAL_2025 = {
    "calibration_year": 2025,
    "description": "2025 parameter set (baseline).",
    "sigma_premium": 0.100,
    "sigma_reserve": 0.090,
    "premium_reserve_corr": 0.50,
    "ir_shock": 0.011,
    "cat_factor": 0.120,
    "op_factor": 0.030,
    "op_cap_of_bscr": 0.30,
    "corr_nl_mkt": 0.25,
    "corr_nl_cat": 0.25,
    "corr_mkt_cat": 0.25,
}
CAL_2026 = {
    **CAL_2025,
    "calibration_year": 2026,
    "description": ("2026 update: premium/reserve sigmas up (claims inflation), "
                    "interest-rate shock up, cat factor up, NL-cat correlation "
                    "strengthened."),
    "sigma_premium": 0.113,
    "sigma_reserve": 0.096,
    "ir_shock": 0.0135,
    "cat_factor": 0.130,
    "corr_nl_cat": 0.35,
}


def compute_scr(inp: dict, cal: dict) -> dict:
    """The model. The pyfunc in 02_model and the Excel formulas mirror this."""
    vp, vr = inp["premium_volume"], inp["reserve_volume"]
    sp, sr = cal["sigma_premium"], cal["sigma_reserve"]
    nl = 3.0 * math.sqrt((sp * vp) ** 2
                         + 2.0 * cal["premium_reserve_corr"] * (sp * vp) * (sr * vr)
                         + (sr * vr) ** 2)
    mkt = abs(inp["assets_mv"] * inp["asset_duration"]
              - inp["liabilities_bel"] * inp["liability_duration"]) * cal["ir_shock"]
    cat = cal["cat_factor"] * vp
    bscr = math.sqrt(nl ** 2 + mkt ** 2 + cat ** 2
                     + 2.0 * (cal["corr_nl_mkt"] * nl * mkt
                              + cal["corr_nl_cat"] * nl * cat
                              + cal["corr_mkt_cat"] * mkt * cat))
    op = min(cal["op_factor"] * vp, cal["op_cap_of_bscr"] * bscr)
    return {"scr_nl": round(nl, 4), "scr_mkt": round(mkt, 4),
            "scr_cat": round(cat, 4), "bscr": round(bscr, 4),
            "op_risk": round(op, 4), "scr": round(bscr + op, 4)}


# ---------------------------------------------------------------------------
# Inputs — 100 entities, ENT-001 has the round numbers shown in the workbook
# ---------------------------------------------------------------------------
LOBS = ["Motor", "Home", "CommercialProperty", "Liability", "Marine"]
ENT1 = {"entity_id": "ENT-001", "entity_name": "Entity 001",
        "line_of_business": "Motor", "premium_volume": 120.0,
        "reserve_volume": 340.0, "assets_mv": 780.0, "asset_duration": 4.2,
        "liabilities_bel": 610.0, "liability_duration": 6.8}

rng = random.Random(2026)
entities = [ENT1]
for i in range(2, 101):
    lob = rng.choice(LOBS)
    scale = {"Motor": 1.0, "Home": 0.6, "CommercialProperty": 0.8,
             "Liability": 0.7, "Marine": 0.4}[lob]
    vp = round(rng.uniform(20, 300) * scale, 1)
    vr = round(vp * rng.uniform(1.5, 4.5), 1)
    bel = round(vr * rng.uniform(1.4, 2.2), 1)
    assets = round(bel * rng.uniform(1.15, 1.45), 1)
    entities.append({
        "entity_id": f"ENT-{i:03d}", "entity_name": f"Entity {i:03d}",
        "line_of_business": lob, "premium_volume": vp, "reserve_volume": vr,
        "assets_mv": assets, "asset_duration": round(rng.uniform(2.5, 6.0), 1),
        "liabilities_bel": bel, "liability_duration": round(rng.uniform(5.0, 11.0), 1),
    })

with open(os.path.join(DATA, "sf_inputs.csv"), "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(ENT1.keys()))
    w.writeheader()
    w.writerows(entities)
print(f"✓ sf_inputs.csv ({len(entities)} entities)")

for name, cal in [("calibration_2025.json", CAL_2025), ("calibration_2026.json", CAL_2026)]:
    with open(os.path.join(DATA, name), "w") as f:
        json.dump(cal, f, indent=2)
    print(f"✓ {name}")

oracle = {"inputs": ENT1,
          "cal_2025": compute_scr(ENT1, CAL_2025),
          "cal_2026": compute_scr(ENT1, CAL_2026)}
with open(os.path.join(DATA, "expected_entity_001.json"), "w") as f:
    json.dump(oracle, f, indent=2)
print("✓ expected_entity_001.json")
print("  ENT-001 SCR @2025:", oracle["cal_2025"]["scr"], " @2026:", oracle["cal_2026"]["scr"],
      f" (+{100*(oracle['cal_2026']['scr']/oracle['cal_2025']['scr']-1):.1f}%)")

# ---------------------------------------------------------------------------
# The workbook — live formulas, one entity
# ---------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEAD = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1B3139")
TITLE = Font(bold=True, size=14)

wb = Workbook()

wsi = wb.active
wsi.title = "Inputs"
wsi["A1"] = "Entity inputs (£m)"
wsi["A1"].font = TITLE
input_rows = [("Premium volume (Vp)", ENT1["premium_volume"]),
              ("Reserve volume (Vr)", ENT1["reserve_volume"]),
              ("Assets market value", ENT1["assets_mv"]),
              ("Asset duration (yrs)", ENT1["asset_duration"]),
              ("Liabilities BEL", ENT1["liabilities_bel"]),
              ("Liability duration (yrs)", ENT1["liability_duration"])]
for i, (label, v) in enumerate(input_rows, start=3):
    wsi.cell(row=i, column=1, value=label).font = Font(bold=True)
    wsi.cell(row=i, column=2, value=v)
wsi.column_dimensions["A"].width = 28

wsc = wb.create_sheet("Calibration")
wsc["A1"] = "Calibration — 2025 parameter set"
wsc["A1"].font = TITLE
wsc["A2"] = ("When the regulator updates these, the actuary retypes this block "
             "and hopes every downstream formula still points at the right cell.")
wsc["A2"].font = Font(italic=True, color="C0392B")
cal_rows = [("sigma_premium", CAL_2025["sigma_premium"]),
            ("sigma_reserve", CAL_2025["sigma_reserve"]),
            ("premium_reserve_corr", CAL_2025["premium_reserve_corr"]),
            ("ir_shock", CAL_2025["ir_shock"]),
            ("cat_factor", CAL_2025["cat_factor"]),
            ("op_factor", CAL_2025["op_factor"]),
            ("op_cap_of_bscr", CAL_2025["op_cap_of_bscr"]),
            ("corr_nl_mkt", CAL_2025["corr_nl_mkt"]),
            ("corr_nl_cat", CAL_2025["corr_nl_cat"]),
            ("corr_mkt_cat", CAL_2025["corr_mkt_cat"])]
for i, (label, v) in enumerate(cal_rows, start=4):
    wsc.cell(row=i, column=1, value=label).font = Font(bold=True)
    wsc.cell(row=i, column=2, value=v)
wsc.column_dimensions["A"].width = 26

# cell references used in the Model tab
I = {k: f"Inputs!B{i}" for i, (k, _) in enumerate(
    [("vp", 0), ("vr", 0), ("assets", 0), ("dur_a", 0), ("bel", 0), ("dur_l", 0)], start=3)}
C = {k: f"Calibration!B{i}" for i, (k, _) in enumerate(cal_rows, start=4)}

wsm = wb.create_sheet("Model")
wsm["A1"] = "Standard Formula — module charges (£m)"
wsm["A1"].font = TITLE
wsm["A3"] = "SCR non-life (prem & reserve)"
wsm["B3"] = (f"=3*SQRT(({C['sigma_premium']}*{I['vp']})^2"
             f"+2*{C['premium_reserve_corr']}*({C['sigma_premium']}*{I['vp']})*({C['sigma_reserve']}*{I['vr']})"
             f"+({C['sigma_reserve']}*{I['vr']})^2)")
wsm["A4"] = "SCR market (interest rate)"
wsm["B4"] = f"=ABS({I['assets']}*{I['dur_a']}-{I['bel']}*{I['dur_l']})*{C['ir_shock']}"
wsm["A5"] = "SCR catastrophe"
wsm["B5"] = f"={C['cat_factor']}*{I['vp']}"
wsm["A7"] = "BSCR (correlation aggregation)"
wsm["B7"] = (f"=SQRT(B3^2+B4^2+B5^2"
             f"+2*({C['corr_nl_mkt']}*B3*B4+{C['corr_nl_cat']}*B3*B5+{C['corr_mkt_cat']}*B4*B5))")
wsm["A8"] = "Operational risk"
wsm["B8"] = f"=MIN({C['op_factor']}*{I['vp']},{C['op_cap_of_bscr']}*B7)"
wsm["A10"] = "SCR"
wsm["A10"].font = Font(bold=True, size=12)
wsm["B10"] = "=B7+B8"
wsm["B10"].font = Font(bold=True, size=12)
for r in list(range(3, 6)) + [7, 8, 10]:
    wsm.cell(row=r, column=2).number_format = "#,##0.00"
wsm.column_dimensions["A"].width = 34

wso = wb.create_sheet("Notes")
wso["A1"] = "About this workbook"
wso["A1"].font = TITLE
for i, line in enumerate([
    "",
    "A deliberately simple Standard Formula model: three risk modules,",
    "a correlation aggregation and an operational-risk add-on, for ONE entity.",
    "",
    "Pain points this file carries in real life:",
    "  • one workbook per entity — a group of 100 means 100 files",
    "  • the calibration is a typed-in block — updates are retyped by hand",
    "  • no version history: 'which parameters produced the Q3 number?'",
    "  • sharing = emailing the file",
    "",
    "Use Case 2 migrates this to a governed, versioned model in Unity Catalog.",
    "All data is synthetic; no customer data is used.",
], start=2):
    wso.cell(row=i, column=1, value=line)
wso.column_dimensions["A"].width = 80

out = os.path.join(HERE, "SF_Model.xlsx")
wb.save(out)
print(f"✓ {out}")
