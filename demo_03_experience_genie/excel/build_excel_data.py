"""Build the "before" workbook: Experience_Monitoring.xlsx + a parity oracle.

This is the *legacy artefact* — the monthly management-information pack an
actuary maintains by hand. It is built from the Excel-sized slice the pipeline
emits (Motor · London · accident year 2024) so its grand totals can be checked
against the Databricks gold table in `05_parity.py`.

Run locally after `01_generate_data.py` has written the slice files to the
Volume and you've copied them into `demo_03_experience_genie/data/`:

    databricks fs cp dbfs:/Volumes/<cat>/actuarial_excel_demo/exp_landing/experience_excel_claims.csv  ../data/
    databricks fs cp dbfs:/Volumes/<cat>/actuarial_excel_demo/exp_landing/experience_excel_premium.csv ../data/
    uv run --with pandas --with openpyxl python build_excel_data.py

Outputs (committed so the repo is self-contained):
    Experience_Monitoring.xlsx   the legacy workbook (Data, Lookup, Pivot, Dashboard tabs)
    parity_oracle.json           grand totals the pivot displays, for 05_parity.py
"""
import json
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "data")

claims = pd.read_csv(os.path.join(DATA, "experience_excel_claims.csv"))
premium = pd.read_csv(os.path.join(DATA, "experience_excel_premium.csv"))

# ----------------------------------------------------------------------------
# The pivot the actuary builds: by channel, earned premium / incurred / loss ratio
# ----------------------------------------------------------------------------
claims["paid"] = claims.apply(
    lambda r: r.transaction_amount if r.transaction_type in ("PAYMENT", "RECOVERY") else 0.0, axis=1)
claims["osr"] = claims.apply(
    lambda r: r.transaction_amount if r.transaction_type == "RESERVE_CHANGE" else 0.0, axis=1)

claims_by_ch = claims.groupby("channel").agg(
    reported_claims=("claim_id", "nunique"),
    paid=("paid", "sum"),
    outstanding=("osr", "sum"),
).reset_index()
claims_by_ch["incurred"] = claims_by_ch.paid + claims_by_ch.outstanding

prem_by_ch = premium.groupby("channel").earned_premium.sum().reset_index()

pivot = prem_by_ch.merge(claims_by_ch, on="channel", how="left").fillna(0)
pivot["loss_ratio"] = pivot.incurred / pivot.earned_premium
pivot = pivot[["channel", "earned_premium", "reported_claims", "paid",
               "outstanding", "incurred", "loss_ratio"]].round(2)

grand = {
    "scope": "Motor / London / accident year 2024",
    "earned_premium": round(float(pivot.earned_premium.sum()), 2),
    "reported_claims": int(claims.claim_id.nunique()),
    "paid": round(float(pivot.paid.sum()), 2),
    "outstanding": round(float(pivot.outstanding.sum()), 2),
    "incurred": round(float(pivot.incurred.sum()), 2),
    "loss_ratio": round(float(pivot.incurred.sum() / pivot.earned_premium.sum()), 4),
}
print("Grand totals (the oracle):")
print(json.dumps(grand, indent=2))

# ----------------------------------------------------------------------------
# Build the workbook
# ----------------------------------------------------------------------------
HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1B3139")  # Databricks navy
TITLE = Font(bold=True, size=14)

wb = Workbook()

# --- Pivot_Experience (what the actuary screenshots into the board pack) ---
ws = wb.active
ws.title = "Pivot_Experience"
ws["A1"] = "Experience Monitoring — Motor · London · AY2024"
ws["A1"].font = TITLE
ws["A2"] = "Source: monthly claims + premium export, pasted into Data tabs, VLOOKUP to Lookup tab."
ws["A3"] = "REFRESH MONTHLY: paste new export, right-click PivotTable → Refresh. ~half a day."
ws["A3"].font = Font(italic=True, color="C0392B")

start = 5
cols = ["channel", "earned_premium", "reported_claims", "paid", "outstanding", "incurred", "loss_ratio"]
for j, c in enumerate(cols, start=1):
    cell = ws.cell(row=start, column=j, value=c)
    cell.font = HEAD
    cell.fill = HEAD_FILL
for i, (_, r) in enumerate(pivot.iterrows(), start=start + 1):
    for j, c in enumerate(cols, start=1):
        v = r[c]
        cell = ws.cell(row=i, column=j, value=v)
        if c == "loss_ratio":
            cell.number_format = "0.0%"
        elif c not in ("channel", "reported_claims"):
            cell.number_format = "#,##0"
gr = start + 1 + len(pivot)
ws.cell(row=gr, column=1, value="Grand Total").font = Font(bold=True)
ws.cell(row=gr, column=2, value=grand["earned_premium"]).number_format = "#,##0"
ws.cell(row=gr, column=3, value=grand["reported_claims"])
ws.cell(row=gr, column=4, value=grand["paid"]).number_format = "#,##0"
ws.cell(row=gr, column=5, value=grand["outstanding"]).number_format = "#,##0"
ws.cell(row=gr, column=6, value=grand["incurred"]).number_format = "#,##0"
lr_cell = ws.cell(row=gr, column=7, value=grand["loss_ratio"])
lr_cell.number_format = "0.0%"
lr_cell.font = Font(bold=True)
for col in "ABCDEFG":
    ws.column_dimensions[col].width = 16

chart = BarChart()
chart.title = "Loss ratio by channel"
chart.type = "col"
data = Reference(ws, min_col=7, min_row=start, max_row=start + len(pivot))
cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(pivot))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "I5")

# --- Data_Claims (the pasted export) ---
wc = wb.create_sheet("Data_Claims")
for r in dataframe_to_rows(claims.drop(columns=["paid", "osr"]), index=False, header=True):
    wc.append(r)
for cell in wc[1]:
    cell.font = HEAD
    cell.fill = HEAD_FILL

# --- Data_Premium ---
wp = wb.create_sheet("Data_Premium")
for r in dataframe_to_rows(premium, index=False, header=True):
    wp.append(r)
for cell in wp[1]:
    cell.font = HEAD
    cell.fill = HEAD_FILL

# --- Lookup (the VLOOKUP tab) ---
wl = wb.create_sheet("Lookup")
seg = claims[["policy_segment", "line_of_business", "region", "channel"]].drop_duplicates()
for r in dataframe_to_rows(seg, index=False, header=True):
    wl.append(r)
for cell in wl[1]:
    cell.font = HEAD
    cell.fill = HEAD_FILL

# --- Dashboard (the board-pack note) ---
wd = wb.create_sheet("Dashboard")
wd["A1"] = "Quarterly Board Pack — Motor London"
wd["A1"].font = TITLE
notes = [
    "",
    "This tab is screenshotted into PowerPoint each quarter.",
    "",
    "Pain points an actuary will recognise:",
    "  • Refresh is manual — paste, VLOOKUP fill-down, refresh every pivot.",
    "  • Only one slice fits: the full multi-year, multi-line book is too big for Excel.",
    "  • Every 'can you also show me X by Y' is a new pivot built by hand.",
    "  • No lineage: which export produced this number? Nobody can say.",
    "",
    "The migration replaces this tab with an AI/BI dashboard (always current) and",
    "a Genie space (ask 'loss ratio for Motor 2024 by channel' in plain English).",
]
for i, line in enumerate(notes, start=2):
    wd.cell(row=i, column=1, value=line)
wd.column_dimensions["A"].width = 90

out_xlsx = os.path.join(HERE, "Experience_Monitoring.xlsx")
wb.save(out_xlsx)
print(f"✓ {out_xlsx}")

with open(os.path.join(HERE, "parity_oracle.json"), "w") as f:
    json.dump(grand, f, indent=2)
print(f"✓ {os.path.join(HERE, 'parity_oracle.json')}")
