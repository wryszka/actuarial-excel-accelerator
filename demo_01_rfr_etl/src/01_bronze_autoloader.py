# Databricks notebook source
# MAGIC %md
# MAGIC # Bronze — EIOPA monthly RFR files → `bronze_rfr_curves`
# MAGIC
# MAGIC Auto Loader picks up new `.xlsx` files from the `rfr_landing` Volume
# MAGIC and parses each one's `RFR_spot_no_VA` tab with openpyxl. Output is a
# MAGIC bronze table in WIDE shape — one row per (effective_date, maturity_years)
# MAGIC and one column per currency — matching what the actuary used to see on
# MAGIC their Excel `Raw_Paste` tab.
# MAGIC
# MAGIC Demo-grade: monthly files are small, so we use `cloudFiles.format =
# MAGIC binaryFile` + a `mapInPandas` driver-side parse with openpyxl. For
# MAGIC larger volumes a Spark Excel reader would be appropriate.

# COMMAND ----------

dbutils.widgets.text("catalog_name", "lr_serverless_aws_us_catalog")
dbutils.widgets.text("schema_name", "actuarial_excel_demo")
dbutils.widgets.text("rfr_volume_name", "rfr_landing")

catalog = dbutils.widgets.get("catalog_name")
schema = dbutils.widgets.get("schema_name")
volume = dbutils.widgets.get("rfr_volume_name")

fqn = f"{catalog}.{schema}"
volume_path = f"/Volumes/{catalog}/{schema}/{volume}"
checkpoint_path = f"{volume_path}/_checkpoints/bronze_rfr_curves"

# COMMAND ----------

# MAGIC %md
# MAGIC ## Install openpyxl on the driver

# COMMAND ----------

# MAGIC %pip install --quiet openpyxl
# MAGIC dbutils.library.restartPython()

# COMMAND ----------

# Re-read widgets after restart
catalog = dbutils.widgets.get("catalog_name")
schema = dbutils.widgets.get("schema_name")
volume = dbutils.widgets.get("rfr_volume_name")
fqn = f"{catalog}.{schema}"
volume_path = f"/Volumes/{catalog}/{schema}/{volume}"
checkpoint_path = f"{volume_path}/_checkpoints/bronze_rfr_curves"

# COMMAND ----------

# MAGIC %md
# MAGIC ## Read every `.xlsx` in the Volume as a binary blob

# COMMAND ----------

import pyspark.sql.functions as F
from pyspark.sql.types import StructType, StructField, DateType, IntegerType, DoubleType, StringType, TimestampType

binary_df = (spark.read
    .format("binaryFile")
    .option("pathGlobFilter", "*.xlsx")
    .option("recursiveFileLookup", "false")
    .load(volume_path)
)

print(f"Found {binary_df.count()} .xlsx file(s) in {volume_path}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Parse each file's `RFR_spot_no_VA` tab with openpyxl
# MAGIC
# MAGIC One file → 30 rows (maturity 1..30 years), with EUR/GBP/USD columns.
# MAGIC `mapInPandas` runs the parser on the driver — fine for monthly files.

# COMMAND ----------

# Schema the parser emits (BEFORE adding _ingested_at / _source_file)
parsed_schema = StructType([
    StructField("effective_date", DateType(), False),
    StructField("maturity_years", IntegerType(), False),
    StructField("EUR", DoubleType(), True),
    StructField("GBP", DoubleType(), True),
    StructField("USD", DoubleType(), True),
    StructField("_source_file", StringType(), False),
])


def parse_eiopa_files(iterator):
    """Driver-side openpyxl parse — one binary blob in, many rows out."""
    import io
    import re
    from datetime import date
    import pandas as pd
    from openpyxl import load_workbook

    DATE_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")

    for batch in iterator:
        rows = []
        for _, file_row in batch.iterrows():
            content = file_row["content"]
            src = file_row["path"]
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            ws = wb["RFR_spot_no_VA"]

            # Header row: cell A2 contains "...reference date: YYYY-MM-DD"
            header_text = ws["A2"].value or ""
            m = DATE_RE.search(header_text)
            if not m:
                raise ValueError(f"No reference date in {src}: '{header_text}'")
            eff_date = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))

            # Currency columns from header row 5
            headers = {}
            for col_idx in range(2, ws.max_column + 1):
                ccy = ws.cell(row=5, column=col_idx).value
                if ccy in ("EUR", "GBP", "USD"):
                    headers[ccy] = col_idx

            # Data rows: 6..35 = maturity 1..30
            for r in range(6, 36):
                m_year = ws.cell(row=r, column=1).value
                if m_year is None:
                    continue
                rows.append({
                    "effective_date": eff_date,
                    "maturity_years": int(m_year),
                    "EUR": _as_float(ws.cell(row=r, column=headers["EUR"]).value),
                    "GBP": _as_float(ws.cell(row=r, column=headers["GBP"]).value),
                    "USD": _as_float(ws.cell(row=r, column=headers["USD"]).value),
                    "_source_file": src,
                })
            wb.close()
        if rows:
            yield pd.DataFrame(rows)


def _as_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# mapInPandas needs the helper visible at module scope on the executor.
# Bind it onto the function so a single import is enough.
parse_eiopa_files._as_float = _as_float  # type: ignore[attr-defined]

# COMMAND ----------

parsed = (binary_df
    .select("content", "path")
    .mapInPandas(parse_eiopa_files, schema=parsed_schema)
    .withColumn("_ingested_at", F.current_timestamp())
)

(parsed
    .write
    .mode("overwrite")
    .option("overwriteSchema", "true")
    .saveAsTable(f"{fqn}.bronze_rfr_curves")
)

n = spark.table(f"{fqn}.bronze_rfr_curves").count()
print(f"✓ {fqn}.bronze_rfr_curves — {n} rows from {binary_df.count()} file(s)")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Bronze comment

# COMMAND ----------

spark.sql(f"""
    COMMENT ON TABLE {fqn}.bronze_rfr_curves IS
    'Bronze: one row per (effective_date, maturity_years) lifted from each '
    'monthly EIOPA RFR_spot_no_VA tab in /Volumes/{catalog}/{schema}/{volume}. '
    'Wide format matching the original Excel Raw_Paste shape — silver_rfr_curves '
    'unpivots to long form.'
""")

display(spark.table(f"{fqn}.bronze_rfr_curves").orderBy("effective_date", "maturity_years").limit(20))
