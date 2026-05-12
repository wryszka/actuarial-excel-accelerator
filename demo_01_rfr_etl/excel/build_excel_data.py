"""
Build the Excel fixtures for demo 1.

This script generates two things:

  1.  ../sample_data/EIOPA_RFR_YYYY_MM.xlsx        (4 monthly files)
      Synthetic EIOPA-shaped monthly publications. These are the demo's
      *inputs* — what you'd download from the EIOPA website. The Databricks
      pipeline (01_bronze_autoloader.py) reads these from the rfr_landing
      Volume.

  2.  ./RFR_Master.xlsx                            (1 fixture)
      An .xlsx version of the macro-driven RFR_Master.xlsm. No VBA, but the
      Raw_Paste / Transform / History tabs are populated with three months
      of synthetic data — enough to exercise the ingestion logic against
      "what Excel would have produced" if the macros ran.

Run once:

    uv run --with openpyxl python demo_01_rfr_etl/excel/build_excel_data.py

The synthetic curves are derived from a deterministic Nelson-Siegel-ish
function with three currency-specific parameter sets. Seeded by year+month
so they're reproducible across runs.
"""

from __future__ import annotations

import math
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
SAMPLE_DIR = HERE.parent / "sample_data"
MASTER_PATH = HERE / "RFR_Master.xlsx"

CURRENCIES = ["EUR", "GBP", "USD"]
MATURITIES_YEARS = list(range(1, 31))  # 1..30 years

# Nelson-Siegel-ish parameters per currency (level, slope, curvature, tau).
# Tweaked to produce plausible spot curves in [-1%, 6%] range.
NS_PARAMS = {
    "EUR": (0.0250, -0.0180, 0.0120, 3.5),
    "GBP": (0.0350, -0.0120, 0.0100, 3.0),
    "USD": (0.0420, -0.0080, 0.0060, 4.0),
}

# The four monthly publications we generate as sample inputs.
PUBLICATIONS = [
    (date(2025, 9, 30), "EIOPA_RFR_2025_09.xlsx"),
    (date(2025, 10, 31), "EIOPA_RFR_2025_10.xlsx"),
    (date(2025, 11, 30), "EIOPA_RFR_2025_11.xlsx"),
    (date(2025, 12, 31), "EIOPA_RFR_2025_12.xlsx"),
]


def nelson_siegel(maturity_years: int, level: float, slope: float,
                  curvature: float, tau: float) -> float:
    """Standard Nelson-Siegel spot rate function."""
    if maturity_years == 0:
        return level + slope
    x = maturity_years / tau
    factor = (1 - math.exp(-x)) / x
    return level + slope * factor + curvature * (factor - math.exp(-x))


def curve(currency: str, effective_date: date) -> dict[int, float]:
    """Deterministic spot curve for (currency, effective_date)."""
    level, slope, curvature, tau = NS_PARAMS[currency]
    # Drift parameters slightly month-to-month so successive curves differ.
    months_since_epoch = (effective_date.year - 2025) * 12 + effective_date.month
    drift = math.sin(months_since_epoch / 6.0) * 0.0015
    level += drift
    return {
        m: round(
            nelson_siegel(m, level, slope, curvature, tau), 6
        )
        for m in MATURITIES_YEARS
    }


# --------------------------------------------------------------------------- #
# Synthetic EIOPA monthly publication
# --------------------------------------------------------------------------- #

def build_eiopa_publication(effective_date: date, out_path: Path) -> None:
    """Write one EIOPA-shaped .xlsx file."""
    wb = Workbook()

    # --- Index tab (placeholder, matches real EIOPA file structure) ---
    idx = wb.active
    idx.title = "Index"
    idx["A1"] = "EIOPA — Risk-Free Interest Rate Term Structures"
    idx["A1"].font = Font(bold=True, size=14)
    idx["A3"] = f"Reference date: {effective_date.isoformat()}"
    idx["A4"] = "Source (synthetic demonstration — not real EIOPA data)"
    idx["A6"] = "Sheets in this workbook:"
    idx["A7"] = "  RFR_spot_no_VA      — spot rates, no volatility adjustment"
    idx["A8"] = "  RFR_spot_with_VA    — spot rates, with volatility adjustment"

    # --- RFR_spot_no_VA — the one our pipeline consumes ---
    no_va = wb.create_sheet("RFR_spot_no_VA")
    _write_rfr_sheet(no_va, effective_date, va_spread_bps=0)

    # --- RFR_spot_with_VA — included for realism only ---
    with_va = wb.create_sheet("RFR_spot_with_VA")
    _write_rfr_sheet(with_va, effective_date, va_spread_bps=12)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


def _write_rfr_sheet(ws, effective_date: date, va_spread_bps: int) -> None:
    """Lay out one of the RFR_spot_* sheets, EIOPA-style."""
    ws["A1"] = "EIOPA Risk-Free Interest Rate Term Structures"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        f"Spot rates - {'with' if va_spread_bps else 'no'} Volatility Adjustment - "
        f"reference date: {effective_date.isoformat()}"
    )
    ws["A3"] = "Synthetic — for demonstration only"
    ws["A3"].font = Font(italic=True, color="808080")

    # Header row at row 5: maturity column + one column per currency
    header_row = 5
    ws.cell(row=header_row, column=1, value="Maturity (in years)").font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(row=header_row, column=1).fill = header_fill
    for j, ccy in enumerate(CURRENCIES, start=2):
        cell = ws.cell(row=header_row, column=j, value=ccy)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # Data rows: one row per maturity, one column per currency
    for i, m in enumerate(MATURITIES_YEARS):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=m)
        for j, ccy in enumerate(CURRENCIES, start=2):
            rate = curve(ccy, effective_date)[m] + va_spread_bps / 10_000
            ws.cell(row=row, column=j, value=round(rate, 6)).number_format = "0.000000"

    # Reasonable column widths
    ws.column_dimensions["A"].width = 22
    for j in range(2, 2 + len(CURRENCIES)):
        ws.column_dimensions[get_column_letter(j)].width = 12


# --------------------------------------------------------------------------- #
# RFR_Master.xlsx — the actuary's workbook, as an .xlsx for testing
# --------------------------------------------------------------------------- #

def build_rfr_master(out_path: Path) -> None:
    """Generate the .xlsx version of RFR_Master.xlsm — same data, no VBA."""
    wb = Workbook()

    # --- Instructions tab ---
    instr = wb.active
    instr.title = "Instructions"
    instr["A1"] = "RFR_Master — Monthly EIOPA Curve Workbook"
    instr["A1"].font = Font(bold=True, size=14)
    instr["A3"] = "This is the .xlsx test fixture. The .xlsm version contains the VBA modules described in VBA_SPEC.md."
    instr["A5"] = "Sheets:"
    instr["A6"] = "  Raw_Paste   — the actuary pastes EIOPA's RFR_spot_no_VA tab here every month"
    instr["A7"] = "  Transform   — ReshapeCurve unpivots maturity × currency into long form"
    instr["A8"] = "  History     — AppendHistory writes each month's curve below the previous one"
    instr["A9"] = "  Chart       — CurveChart binds to a History range; RefreshChart re-binds after append"

    # --- Raw_Paste: latest month, in EIOPA shape ---
    raw = wb.create_sheet("Raw_Paste")
    latest = PUBLICATIONS[-1][0]
    _write_rfr_sheet(raw, latest, va_spread_bps=0)

    # --- Transform: long-form for latest month ---
    tx = wb.create_sheet("Transform")
    tx["A1"] = f"Long-form curve — reference date {latest.isoformat()}"
    tx["A1"].font = Font(bold=True)
    tx["A3"] = "effective_date"
    tx["B3"] = "currency"
    tx["C3"] = "maturity_years"
    tx["D3"] = "spot_rate"
    for c in ["A3", "B3", "C3", "D3"]:
        tx[c].font = Font(bold=True)
    r = 4
    for ccy in CURRENCIES:
        curve_data = curve(ccy, latest)
        for m, rate in curve_data.items():
            tx.cell(row=r, column=1, value=latest.isoformat())
            tx.cell(row=r, column=2, value=ccy)
            tx.cell(row=r, column=3, value=m)
            tx.cell(row=r, column=4, value=rate).number_format = "0.000000"
            r += 1

    # --- History: three months of curves stacked ---
    hist = wb.create_sheet("History")
    hist["A1"] = "Accumulated curve history"
    hist["A1"].font = Font(bold=True)
    hist["A3"] = "effective_date"
    hist["B3"] = "currency"
    hist["C3"] = "maturity_years"
    hist["D3"] = "spot_rate"
    for c in ["A3", "B3", "C3", "D3"]:
        hist[c].font = Font(bold=True)
    r = 4
    for eff_date, _ in PUBLICATIONS[:3]:
        for ccy in CURRENCIES:
            curve_data = curve(ccy, eff_date)
            for m, rate in curve_data.items():
                hist.cell(row=r, column=1, value=eff_date.isoformat())
                hist.cell(row=r, column=2, value=ccy)
                hist.cell(row=r, column=3, value=m)
                hist.cell(row=r, column=4, value=rate).number_format = "0.000000"
                r += 1

    # --- Chart tab (placeholder header — the .xlsm wires the real chart) ---
    ch = wb.create_sheet("Chart")
    ch["A1"] = "Curve viz — bound by RefreshChart() in the .xlsm version"

    for sheet in (instr, raw, tx, hist, ch):
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 16
        sheet.column_dimensions["D"].width = 14

    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


def main() -> None:
    print("Building monthly EIOPA publications:")
    for eff_date, name in PUBLICATIONS:
        build_eiopa_publication(eff_date, SAMPLE_DIR / name)

    print("\nBuilding RFR_Master.xlsx fixture:")
    build_rfr_master(MASTER_PATH)

    print("\nDone.")
    print(f"  Sample inputs : {SAMPLE_DIR}")
    print(f"  Master fixture: {MASTER_PATH}")


if __name__ == "__main__":
    main()
