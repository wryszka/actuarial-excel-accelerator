# Databricks notebook source
# MAGIC %md
# MAGIC # 08 — Parity test (Excel ↔ Databricks)
# MAGIC
# MAGIC Loads the `SCR_Computed` hidden tab from
# MAGIC `demo_02a_scr_sf/excel/SCR_StandardFormula.xlsx`, which carries the
# MAGIC Python-computed reference values written by `build_excel_data.py`.
# MAGIC Then calls the orchestrator with the same base scenario and compares
# MAGIC sub-module by sub-module within a 1% tolerance.
# MAGIC
# MAGIC The visible Excel tabs carry formulas — an actuary can open the file
# MAGIC in Excel, hit F9, and watch the formulas evaluate to the same numbers.
# MAGIC That's the human-side parity check; this notebook is the
# MAGIC machine-readable one.
# MAGIC
# MAGIC **Tolerance is 1% relative or €1k absolute, whichever is larger.**

# COMMAND ----------

# MAGIC %pip install --quiet openpyxl
# MAGIC dbutils.library.restartPython()

# COMMAND ----------

dbutils.widgets.text("catalog_name", "lr_dev_aws_us_catalog")
dbutils.widgets.text("schema_name", "actuarial_excel_demo")

catalog = dbutils.widgets.get("catalog_name")
schema = dbutils.widgets.get("schema_name")
fqn = f"{catalog}.{schema}"

# COMMAND ----------

# MAGIC %md
# MAGIC ## Locate the .xlsx (bundled next to this notebook)

# COMMAND ----------

import os
notebook_path = (
    dbutils.notebook.entry_point.getDbutils().notebook().getContext()
    .notebookPath().getOrElse(None)
)
notebook_dir = os.path.dirname(notebook_path) if notebook_path else "."
xlsx_path = f"/Workspace{notebook_dir}/../excel/SCR_StandardFormula.xlsx"
print(f"Loading parity oracle from {xlsx_path}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Read the SCR_Computed hidden tab

# COMMAND ----------

from openpyxl import load_workbook

wb = load_workbook(xlsx_path, data_only=False)
ws = wb["SCR_Computed"]

oracle = {}
for r in range(4, 11):
    key = ws.cell(row=r, column=1).value
    val = ws.cell(row=r, column=2).value
    if key:
        oracle[key] = float(val)
print("Excel oracle (Python-computed reference values in SCR_Computed):")
for k, v in oracle.items():
    print(f"  {k:18s} {v:>20,.2f}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Compute the Databricks side
# MAGIC
# MAGIC Inlines the orchestrator (same body as 06_orchestrator.py) so this
# MAGIC notebook is self-contained and doesn't need a `%run` of 06.

# COMMAND ----------

import math


def combined_sigma(sp, sr, vp, vr, alpha=0.5):
    v = vp + vr
    if v == 0:
        return 0.0
    return math.sqrt((sp*vp)**2 + 2*alpha*sp*sr*vp*vr + (sr*vr)**2) / v


def compute_databricks(scenario_id="base"):
    inputs_row = spark.table(f"{fqn}.scr_inputs").filter(f"scenario_id = '{scenario_id}'").collect()[0]
    ass_row = spark.table(f"{fqn}.scr_assumptions").filter("is_current = 'true'").collect()[0]
    rfr_rows = (spark.table(f"{fqn}.rfr_curves")
                .filter(f"effective_date = '{inputs_row['rfr_effective_date']}'"
                        f" AND currency = '{inputs_row['currency']}'")
                .select("maturity_months", "spot_rate").collect())

    lob_order = list(ass_row["nl_lob_order"])
    sigmas = {r["lob"]: r for r in ass_row["nl_lob_sigmas"]}
    vols = {r["lob"]: r for r in inputs_row["lob_volumes"]}
    rho_nl = [list(r) for r in ass_row["nl_lob_correlation"]]
    sigma_lob, v_lob = [], []
    for lob in lob_order:
        vp = float(vols[lob]["v_prem"])
        vr = float(vols[lob]["v_res"])
        sigma_lob.append(combined_sigma(
            float(sigmas[lob]["sigma_prem"]), float(sigmas[lob]["sigma_res"]), vp, vr,
        ))
        v_lob.append(vp + vr)
    v_nl = sum(v_lob)
    inner = sum(rho_nl[i][j]*sigma_lob[i]*v_lob[i]*sigma_lob[j]*v_lob[j]
                for i in range(len(lob_order)) for j in range(len(lob_order)))
    sigma_nl = math.sqrt(max(inner, 0.0)) / v_nl
    scr_uw = 3.0 * sigma_nl * v_nl

    # Market IR — use the actual RFR curve from demo 1
    curve = {int(r["maturity_months"]): float(r["spot_rate"]) for r in rfr_rows}
    su = float(ass_row["ir_shock_up_bps"]) / 10_000
    sd = float(ass_row["ir_shock_down_bps"]) / 10_000
    floor = float(ass_row["ir_shock_down_floor"])

    def pv(c):
        return sum(
            float(cf["amount"]) / (1.0 + (c.get(int(cf["year"])*12)
                                          or c[min(c.keys(), key=lambda m: abs(m - int(cf["year"])*12))]))
            ** int(cf["year"])
            for cf in inputs_row["liability_cash_flows"]
        )

    asset = float(inputs_row["asset_value"])
    dur = float(inputs_row["asset_modified_duration"])
    pv_b = pv(curve)
    pv_u = pv({mo: r + su for mo, r in curve.items()})
    pv_d = pv({mo: max(r + sd, floor) for mo, r in curve.items()})
    nav_b = asset - pv_b
    nav_u = asset * (1 - dur * su) - pv_u
    nav_d = asset * (1 - dur * sd) - pv_d
    scr_mkt = max(nav_b - min(nav_u, nav_d), 0.0)

    scr_cat = float(ass_row["cat_plug"])
    rho = float(ass_row["bscr_rho_market_uw"])
    bscr = math.sqrt(scr_uw**2 + scr_mkt**2 + 2*rho*scr_uw*scr_mkt) + scr_cat
    op = float(ass_row["op_factor"]) * float(inputs_row["earned_premium"])
    lacdt = float(ass_row["lacdt"])
    scr = bscr + op - lacdt
    return {
        "scr_nl_premres": scr_uw, "scr_mkt_ir": scr_mkt, "scr_cat": scr_cat,
        "bscr": bscr, "op_risk": op, "lacdt": lacdt, "scr": scr,
    }


# COMMAND ----------

# MAGIC %md
# MAGIC ## Note on `scr_mkt_ir` parity
# MAGIC
# MAGIC The Excel side uses a flat 2.5% fallback rate (the `Market_IR` tab's
# MAGIC `B3` cell) because the workbook is committed without a curve paste.
# MAGIC The Python oracle in `build_excel_data.py` also uses 2.5%. The
# MAGIC Databricks side reads the real RFR curve from demo 1's `rfr_curves`.
# MAGIC
# MAGIC We therefore expect `scr_mkt_ir` to differ between Excel oracle and
# MAGIC Databricks. The parity test reports both numbers and flags the row,
# MAGIC but only fails on differences in `scr_nl_premres`, `scr_cat`,
# MAGIC `op_risk`, `lacdt` — the four lines that are curve-independent.

# COMMAND ----------

dbx = compute_databricks("base")
print("Databricks compute:")
for k, v in dbx.items():
    print(f"  {k:18s} {v:>20,.2f}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Compare

# COMMAND ----------

REL_TOL = 0.01
ABS_TOL = 1_000.0
CURVE_INDEPENDENT = {"scr_nl_premres", "scr_cat", "op_risk", "lacdt"}

failures = []
rows = []
for k in ["scr_nl_premres", "scr_mkt_ir", "scr_cat", "bscr", "op_risk", "lacdt", "scr"]:
    a, b = oracle[k], dbx[k]
    diff = b - a
    rel = abs(diff) / max(abs(a), 1.0)
    inside = (abs(diff) <= ABS_TOL) or (rel <= REL_TOL)
    if k in CURVE_INDEPENDENT and not inside:
        failures.append(f"{k}: oracle={a:,.2f} dbx={b:,.2f} diff={diff:,.2f} ({rel:.2%})")
    rows.append({
        "metric": k,
        "excel_oracle": round(a, 2),
        "databricks": round(b, 2),
        "diff": round(diff, 2),
        "rel_pct": round(rel * 100, 3),
        "curve_dep": k not in CURVE_INDEPENDENT,
        "verdict": "OK" if inside else ("(curve mismatch)" if k not in CURVE_INDEPENDENT else "FAIL"),
    })

import pandas as pd
display(spark.createDataFrame(pd.DataFrame(rows)))

if failures:
    print("\nFAIL — curve-independent metrics drifted:")
    for f in failures:
        print(f"  - {f}")
    dbutils.notebook.exit("FAIL: " + "; ".join(failures))
print("\nOK — curve-independent metrics within tolerance.")
dbutils.notebook.exit("OK")
