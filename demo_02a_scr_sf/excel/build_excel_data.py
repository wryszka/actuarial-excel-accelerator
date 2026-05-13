"""
Build the Excel fixture for demo 2A: the SCR Standard Formula workbook.

Produces ./SCR_StandardFormula.xlsx — a no-VBA .xlsx version of the
macro-driven SCR_StandardFormula.xlsm. The .xlsm is hand-authored from
VBA_SPEC.md; this script generates the .xlsx so that:

  1.  Anyone (including CI, the parity test, or a reader without macros
      enabled) can open the workbook and see the full model.

  2.  The parity test in src/08_parity_test.py has a deterministic
      reference: a hidden `SCR_Computed` tab holds Python-computed
      values that the Databricks orchestrator must match.

The Excel-side formulas in the visible tabs (Assumptions, NL_PremRes,
Market_IR, Aggregation) are written out so an actuary opening the
workbook in Excel can press F9 and see Excel evaluate them to the
same numbers — but the parity test does not rely on Excel
evaluation; it reads the literal values in `SCR_Computed`.

Run once:

    uv run --with openpyxl python demo_02a_scr_sf/excel/build_excel_data.py
"""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DEMO_ROOT = HERE.parent
SAMPLE = DEMO_ROOT / "sample_data"
OUT_PATH = HERE / "SCR_StandardFormula.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
HIDDEN_NOTE_FILL = PatternFill("solid", fgColor="FCE4D6")


# --------------------------------------------------------------------------- #
# Pure-Python reference compute — mirrors the Excel formulas and the
# Databricks orchestrator. This is the parity oracle.
# --------------------------------------------------------------------------- #

def combined_sigma(sigma_prem: float, sigma_res: float,
                   v_prem: float, v_res: float, alpha: float = 0.5) -> float:
    """EIOPA-style combined sigma over premium + reserve with α=0.5."""
    v = v_prem + v_res
    if v == 0:
        return 0.0
    num = (
        (sigma_prem * v_prem) ** 2
        + 2 * alpha * sigma_prem * sigma_res * v_prem * v_res
        + (sigma_res * v_res) ** 2
    )
    return math.sqrt(num) / v


def nl_premres_scr(inputs: dict, assumptions: dict,
                   shock_uplifts: dict[str, float] | None = None) -> float:
    """Aggregate NL Premium & Reserve SCR across the 4 LoBs.

    shock_uplifts is e.g. {"Motor": 0.10} → 10% uplift on Motor volumes.
    """
    shock_uplifts = shock_uplifts or {}
    lob_order: list[str] = assumptions["nl_lob_order"]
    sigmas = {row["lob"]: row for row in assumptions["nl_lob_sigmas"]}
    vols = {row["lob"]: row for row in inputs["lob_volumes"]}
    rho = assumptions["nl_lob_correlation"]

    sigma_lob = []
    v_lob = []
    for lob in lob_order:
        uplift = 1.0 + shock_uplifts.get(lob, 0.0)
        v_prem = vols[lob]["v_prem"] * uplift
        v_res = vols[lob]["v_res"] * uplift
        sigma_lob.append(combined_sigma(
            sigmas[lob]["sigma_prem"], sigmas[lob]["sigma_res"],
            v_prem, v_res,
        ))
        v_lob.append(v_prem + v_res)

    v_nl = sum(v_lob)
    if v_nl == 0:
        return 0.0
    # σ_NL = sqrt( Σ ρ_ij · (σ_i V_i) · (σ_j V_j) ) / V_NL
    inner = 0.0
    for i in range(len(lob_order)):
        for j in range(len(lob_order)):
            inner += rho[i][j] * sigma_lob[i] * v_lob[i] * sigma_lob[j] * v_lob[j]
    sigma_nl = math.sqrt(max(inner, 0.0)) / v_nl
    return 3.0 * sigma_nl * v_nl


def market_ir_scr(inputs: dict, assumptions: dict,
                  ir_shock_bps_override: int | None = None,
                  rfr_curve: dict[int, float] | None = None) -> float:
    """Market IR sub-module — max NAV impact across up/down parallel shocks.

    rfr_curve maps maturity_months → spot_rate. When None, uses a flat 2.5%
    fallback curve (only used by the Excel populator for closed-form numbers;
    the Databricks orchestrator always passes a real curve from rfr_curves).
    """
    if rfr_curve is None:
        rfr_curve = {m * 12: 0.025 for m in range(1, 31)}

    shock_up = (ir_shock_bps_override or assumptions["ir_shock_up_bps"]) / 10_000
    shock_dn = assumptions["ir_shock_down_bps"] / 10_000
    floor = assumptions["ir_shock_down_floor"]

    def pv_liab(curve: dict[int, float]) -> float:
        pv = 0.0
        for cf in inputs["liability_cash_flows"]:
            year = cf["year"]
            mo = year * 12
            r = curve.get(mo)
            if r is None:
                # nearest-maturity lookup for sparse curves
                r = curve[min(curve.keys(), key=lambda m: abs(m - mo))]
            pv += cf["amount"] / (1.0 + r) ** year
        return pv

    asset_mv = inputs["asset_value"]
    duration = inputs["asset_modified_duration"]

    def shocked_curve(shock: float, floor_at: float | None = None) -> dict[int, float]:
        return {
            mo: max(r + shock, floor_at) if floor_at is not None else r + shock
            for mo, r in rfr_curve.items()
        }

    pv_base = pv_liab(rfr_curve)
    nav_base = asset_mv - pv_base

    # Assets reval via modified duration: ΔMV = -D · MV · Δy
    asset_up = asset_mv * (1 - duration * shock_up)
    asset_dn = asset_mv * (1 - duration * shock_dn)

    pv_up = pv_liab(shocked_curve(shock_up))
    pv_dn = pv_liab(shocked_curve(shock_dn, floor_at=floor))

    nav_up = asset_up - pv_up
    nav_dn = asset_dn - pv_dn

    return max(nav_base - min(nav_up, nav_dn), 0.0)


def compute_scr_oracle(inputs: dict, assumptions: dict,
                       shocks: dict | None = None) -> dict:
    """Full SCR — returns the breakdown used by the parity test."""
    shocks = shocks or {}
    scr_uw = nl_premres_scr(
        inputs, assumptions,
        shock_uplifts=shocks.get("lob_uplifts", {}),
    )
    scr_mkt = market_ir_scr(
        inputs, assumptions,
        ir_shock_bps_override=shocks.get("ir_shock_bps"),
    )
    scr_cat = float(assumptions["cat_plug"])

    rho = float(assumptions["bscr_rho_market_uw"])
    bscr = math.sqrt(scr_uw ** 2 + scr_mkt ** 2 + 2 * rho * scr_uw * scr_mkt) + scr_cat
    op = float(assumptions["op_factor"]) * float(inputs["earned_premium"])
    lacdt = float(assumptions["lacdt"])
    scr = bscr + op - lacdt
    return {
        "scr_nl_premres": scr_uw,
        "scr_mkt_ir": scr_mkt,
        "scr_cat": scr_cat,
        "bscr": bscr,
        "op_risk": op,
        "lacdt": lacdt,
        "scr": scr,
    }


# --------------------------------------------------------------------------- #
# Excel workbook layout
# --------------------------------------------------------------------------- #

def _h(cell, fill=HEADER_FILL, bold=True):
    cell.font = Font(bold=bold)
    cell.fill = fill


def build_workbook(inputs: dict, assumptions: dict, out_path: Path) -> None:
    wb = Workbook()

    instr = wb.active
    instr.title = "Instructions"
    _build_instructions(instr)

    inp = wb.create_sheet("Inputs")
    _build_inputs(inp, inputs)

    rfr = wb.create_sheet("RFR_Curves")
    _build_rfr_placeholder(rfr)

    ass = wb.create_sheet("Assumptions")
    _build_assumptions(ass, assumptions)

    nl = wb.create_sheet("NL_PremRes")
    _build_nl_premres(nl, inputs, assumptions)

    mkt = wb.create_sheet("Market_IR")
    _build_market_ir(mkt, inputs, assumptions)

    cat = wb.create_sheet("Cat")
    _build_cat(cat)

    agg = wb.create_sheet("Aggregation")
    _build_aggregation(agg, inputs)

    scen = wb.create_sheet("Scenarios")
    _build_scenarios_header(scen)

    out = wb.create_sheet("Output")
    _build_output(out)

    # Parity oracle — values computed in Python. Hidden by default; the
    # parity test reads these cells via openpyxl (no Excel needed).
    computed = wb.create_sheet("SCR_Computed")
    _build_computed(computed, inputs, assumptions)
    computed.sheet_state = "hidden"

    for ws in wb.worksheets:
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(out_path)
    print(f"  ✓ {out_path}")


def _build_instructions(ws):
    ws["A1"] = "SCR_StandardFormula — Demo 2A Workbook"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "This is the .xlsx test fixture. The .xlsm version contains the VBA modules described in VBA_SPEC.md."
    ws["A5"] = "Tabs (in workflow order):"
    ws["A6"] = "  Inputs       — per-LoB volumes, asset value/duration, liability cash flows"
    ws["A7"] = "  RFR_Curves   — pasted from demo 1 gold table {catalog}.{schema}.rfr_curves"
    ws["A8"] = "  Assumptions  — sigmas, correlation matrices, Op factor, Cat plug, LACDT"
    ws["A9"] = "  NL_PremRes   — per-LoB combined sigma and NL aggregation"
    ws["A10"] = "  Market_IR    — asset and liability NAV under IR up/down shock"
    ws["A11"] = "  Cat          — plug pass-through from Assumptions"
    ws["A12"] = "  Aggregation  — BSCR closed form + Op + LACDT → SCR"
    ws["A13"] = "  Scenarios    — populated by VBA RunScenarios()"
    ws["A14"] = "  Output       — formatted SCR breakdown (built by GenerateSummary)"
    ws["A16"] = "Hidden tab SCR_Computed holds Python-computed reference values"
    ws["A17"] = "used by the Databricks parity test (src/08_parity_test.py)."


def _build_inputs(ws, inputs: dict):
    ws["A1"] = "Inputs — base scenario"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "Scalar"
    ws["B3"] = "Value"
    _h(ws["A3"]); _h(ws["B3"])
    ws["A4"] = "scenario_id"
    ws["B4"] = inputs["scenario_id"]
    ws["A5"] = "as_of_date"
    ws["B5"] = inputs["as_of_date"]
    ws["A6"] = "currency"
    ws["B6"] = inputs["currency"]
    ws["A7"] = "earned_premium"
    ws["B7"] = inputs["earned_premium"]
    ws["A8"] = "asset_value"
    ws["B8"] = inputs["asset_value"]
    ws["A9"] = "asset_modified_duration"
    ws["B9"] = inputs["asset_modified_duration"]

    ws["A11"] = "LoB"
    ws["B11"] = "V_prem"
    ws["C11"] = "V_res"
    for c in ["A11", "B11", "C11"]:
        _h(ws[c])
    for i, row in enumerate(inputs["lob_volumes"], start=12):
        ws.cell(row=i, column=1, value=row["lob"])
        ws.cell(row=i, column=2, value=row["v_prem"]).number_format = "#,##0"
        ws.cell(row=i, column=3, value=row["v_res"]).number_format = "#,##0"

    ws["E11"] = "Year"
    ws["F11"] = "Liability CF"
    for c in ["E11", "F11"]:
        _h(ws[c])
    for i, row in enumerate(inputs["liability_cash_flows"], start=12):
        ws.cell(row=i, column=5, value=row["year"])
        ws.cell(row=i, column=6, value=row["amount"]).number_format = "#,##0"


def _build_rfr_placeholder(ws):
    ws["A1"] = "RFR_Curves — paste from demo 1 gold table"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "Place the rfr_curves rows here (effective_date, currency, maturity_months, spot_rate)."
    ws["A4"] = "The .xlsm version uses the actuary's manual paste from a Databricks SQL query."
    ws["A5"] = "Demo 2A's Excel parity reference does not depend on this tab — it embeds a flat 2.5%"
    ws["A6"] = "fallback so the workbook always evaluates. The Databricks orchestrator uses the real curve."


def _build_assumptions(ws, ass: dict):
    ws["A1"] = "Assumptions — " + ass["assumption_version"]
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "LoB"
    ws["B3"] = "σ_prem"
    ws["C3"] = "σ_res"
    for c in ["A3", "B3", "C3"]:
        _h(ws[c])
    for i, row in enumerate(ass["nl_lob_sigmas"], start=4):
        ws.cell(row=i, column=1, value=row["lob"])
        ws.cell(row=i, column=2, value=row["sigma_prem"]).number_format = "0.0000"
        ws.cell(row=i, column=3, value=row["sigma_res"]).number_format = "0.0000"

    ws["E3"] = "NL LoB correlation matrix"
    _h(ws["E3"])
    for j, lob in enumerate(ass["nl_lob_order"], start=2):
        ws.cell(row=4, column=4 + j, value=lob).font = Font(bold=True)
    for i, lob_i in enumerate(ass["nl_lob_order"], start=5):
        ws.cell(row=i, column=5, value=lob_i).font = Font(bold=True)
        for j, _ in enumerate(ass["nl_lob_order"], start=6):
            ws.cell(row=i, column=j, value=ass["nl_lob_correlation"][i - 5][j - 6])\
                .number_format = "0.00"

    ws["A10"] = "BSCR ρ(Market, UW)"
    ws["B10"] = ass["bscr_rho_market_uw"]
    ws["A11"] = "Op factor"
    ws["B11"] = ass["op_factor"]
    ws["A12"] = "Cat plug"
    ws["B12"] = ass["cat_plug"]
    ws["B12"].number_format = "#,##0"
    ws["A13"] = "LACDT"
    ws["B13"] = ass["lacdt"]
    ws["B13"].number_format = "#,##0"
    ws["A14"] = "IR shock up (bps)"
    ws["B14"] = ass["ir_shock_up_bps"]
    ws["A15"] = "IR shock down (bps)"
    ws["B15"] = ass["ir_shock_down_bps"]
    ws["A16"] = "IR down floor"
    ws["B16"] = ass["ir_shock_down_floor"]


def _build_nl_premres(ws, inputs: dict, ass: dict):
    ws["A1"] = "NL Premium & Reserve — per-LoB and aggregation"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "LoB"
    ws["B3"] = "V_prem"
    ws["C3"] = "V_res"
    ws["D3"] = "σ_prem"
    ws["E3"] = "σ_res"
    ws["F3"] = "V_lob"
    ws["G3"] = "σ_lob"
    for c in ["A3", "B3", "C3", "D3", "E3", "F3", "G3"]:
        _h(ws[c])

    vol_by_lob = {r["lob"]: r for r in inputs["lob_volumes"]}
    sig_by_lob = {r["lob"]: r for r in ass["nl_lob_sigmas"]}
    for i, lob in enumerate(ass["nl_lob_order"], start=4):
        v = vol_by_lob[lob]
        s = sig_by_lob[lob]
        ws.cell(row=i, column=1, value=lob)
        ws.cell(row=i, column=2, value=v["v_prem"]).number_format = "#,##0"
        ws.cell(row=i, column=3, value=v["v_res"]).number_format = "#,##0"
        ws.cell(row=i, column=4, value=s["sigma_prem"]).number_format = "0.0000"
        ws.cell(row=i, column=5, value=s["sigma_res"]).number_format = "0.0000"
        ws.cell(row=i, column=6, value=f"=B{i}+C{i}").number_format = "#,##0"
        # combined sigma EIOPA-shape with α=0.5
        ws.cell(
            row=i, column=7,
            value=(
                f"=IFERROR(SQRT((D{i}*B{i})^2+2*0.5*D{i}*E{i}*B{i}*C{i}+"
                f"(E{i}*C{i})^2)/F{i},0)"
            ),
        ).number_format = "0.0000"

    # Aggregation block below the per-LoB rows.
    ws["A9"] = "Aggregation"
    _h(ws["A9"], fill=TOTAL_FILL)
    ws["A10"] = "V_NL"
    ws["B10"] = "=SUM(F4:F7)"
    ws["B10"].number_format = "#,##0"
    ws["A11"] = "σ_NL"
    ws["B11"] = (
        "=SQRT("
        + "+".join(
            f"Assumptions!{get_column_letter(6 + j)}{5 + i}*G{4 + i}*F{4 + i}*G{4 + j}*F{4 + j}"
            for i in range(4) for j in range(4)
        )
        + ")/B10"
    )
    ws["B11"].number_format = "0.0000"
    ws["A12"] = "SCR_nl_premres"
    ws["B12"] = "=3*B11*B10"
    ws["B12"].number_format = "#,##0"
    ws["B12"].fill = TOTAL_FILL


def _build_market_ir(ws, inputs: dict, ass: dict):
    ws["A1"] = "Market IR — NAV revaluation"
    ws["A1"].font = Font(bold=True, size=12)

    # Flat-rate fallback (the .xlsm pulls from RFR_Curves; we keep a closed
    # form here so the file always evaluates without the curve paste).
    ws["A3"] = "Fallback flat rate"
    ws["B3"] = 0.025
    ws["B3"].number_format = "0.0000"

    ws["A5"] = "Shock up (decimal)"
    ws["B5"] = "=Assumptions!B14/10000"
    ws["A6"] = "Shock down (decimal)"
    ws["B6"] = "=Assumptions!B15/10000"

    ws["A8"] = "Asset MV base"
    ws["B8"] = "=Inputs!B8"
    ws["B8"].number_format = "#,##0"
    ws["A9"] = "Modified duration"
    ws["B9"] = "=Inputs!B9"

    # PV of liability CFs at the flat rate
    ws["A11"] = "PV liabilities (base)"
    ws["B11"] = (
        "=SUMPRODUCT(Inputs!F12:F29/(1+B3)^Inputs!E12:E29)"
    )
    ws["B11"].number_format = "#,##0"

    ws["A12"] = "PV liabilities (IR up)"
    ws["B12"] = (
        "=SUMPRODUCT(Inputs!F12:F29/(1+B3+B5)^Inputs!E12:E29)"
    )
    ws["B12"].number_format = "#,##0"

    ws["A13"] = "PV liabilities (IR dn)"
    ws["B13"] = (
        "=SUMPRODUCT(Inputs!F12:F29/(1+MAX(B3+B6,Assumptions!B16))^Inputs!E12:E29)"
    )
    ws["B13"].number_format = "#,##0"

    ws["A15"] = "Asset MV (IR up)"
    ws["B15"] = "=B8*(1-B9*B5)"
    ws["B15"].number_format = "#,##0"
    ws["A16"] = "Asset MV (IR dn)"
    ws["B16"] = "=B8*(1-B9*B6)"
    ws["B16"].number_format = "#,##0"

    ws["A18"] = "NAV base"
    ws["B18"] = "=B8-B11"
    ws["A19"] = "NAV IR up"
    ws["B19"] = "=B15-B12"
    ws["A20"] = "NAV IR dn"
    ws["B20"] = "=B16-B13"
    for r in (18, 19, 20):
        ws.cell(row=r, column=2).number_format = "#,##0"

    ws["A22"] = "SCR_mkt_ir"
    ws["B22"] = "=MAX(B18-MIN(B19,B20),0)"
    ws["B22"].number_format = "#,##0"
    ws["B22"].fill = TOTAL_FILL


def _build_cat(ws):
    ws["A1"] = "Cat risk — plug pass-through"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "SCR_cat"
    ws["B3"] = "=Assumptions!B12"
    ws["B3"].number_format = "#,##0"
    ws["B3"].fill = TOTAL_FILL
    ws["A5"] = (
        "Demo simplification: Cat is read from Assumptions as a plug. "
        "A real Standard Formula models Cat per peril with reinsurance recoveries."
    )
    ws["A5"].font = Font(italic=True, color="808080")


def _build_aggregation(ws, inputs: dict):
    ws["A1"] = "BSCR aggregation → SCR"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "SCR_uw (NL P&R)"
    ws["B3"] = "=NL_PremRes!B12"
    ws["A4"] = "SCR_mkt (IR)"
    ws["B4"] = "=Market_IR!B22"
    ws["A5"] = "SCR_cat (plug)"
    ws["B5"] = "=Cat!B3"
    ws["A6"] = "ρ (Mkt, UW)"
    ws["B6"] = "=Assumptions!B10"

    ws["A8"] = "BSCR"
    ws["B8"] = "=SQRT(B3^2+B4^2+2*B6*B3*B4)+B5"
    ws["B8"].number_format = "#,##0"
    ws["A9"] = "Op risk"
    ws["B9"] = "=Assumptions!B11*Inputs!B7"
    ws["B9"].number_format = "#,##0"
    ws["A10"] = "LACDT"
    ws["B10"] = "=Assumptions!B13"
    ws["B10"].number_format = "#,##0"

    ws["A12"] = "SCR"
    ws["B12"] = "=B8+B9-B10"
    ws["B12"].number_format = "#,##0"
    ws["A12"].font = Font(bold=True)
    ws["B12"].fill = TOTAL_FILL
    ws["B12"].font = Font(bold=True)


def _build_scenarios_header(ws):
    ws["A1"] = "Scenarios — populated by VBA RunScenarios()"
    ws["A1"].font = Font(bold=True, size=12)
    headers = [
        "scenario_id", "ir_shock_bps",
        "motor_uplift", "property_uplift", "liability_uplift", "other_uplift",
        "scr_nl_premres", "scr_mkt_ir", "scr_cat",
        "bscr", "op_risk", "lacdt", "scr",
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j, value=h)
        _h(ws.cell(row=3, column=j))


def _build_output(ws):
    ws["A1"] = "Output — populated by VBA GenerateSummary()"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "Run GenerateSummary in the .xlsm to populate this tab."


def _build_computed(ws, inputs: dict, ass: dict):
    """Hidden tab with Python-computed reference values for the parity test."""
    oracle = compute_scr_oracle(inputs, ass)
    ws["A1"] = "Parity oracle — Python-computed reference"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        "DO NOT EDIT BY HAND. Regenerated by build_excel_data.py. "
        "src/08_parity_test.py reads cell B<n> for each metric below."
    )
    ws["A2"].fill = HIDDEN_NOTE_FILL

    rows = [
        ("scr_nl_premres", oracle["scr_nl_premres"]),
        ("scr_mkt_ir",     oracle["scr_mkt_ir"]),
        ("scr_cat",        oracle["scr_cat"]),
        ("bscr",           oracle["bscr"]),
        ("op_risk",        oracle["op_risk"]),
        ("lacdt",          oracle["lacdt"]),
        ("scr",            oracle["scr"]),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=float(v))


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #

def main() -> None:
    with open(SAMPLE / "scr_inputs.json") as f:
        inputs_doc = json.load(f)
    with open(SAMPLE / "scr_assumptions.json") as f:
        ass_doc = json.load(f)

    base_inputs = next(s for s in inputs_doc["scenarios"] if s["scenario_id"] == "base")
    current = next(a for a in ass_doc["versions"] if a["is_current"])

    print("Building SCR_StandardFormula.xlsx ...")
    build_workbook(base_inputs, current, OUT_PATH)

    oracle = compute_scr_oracle(base_inputs, current)
    print("\nPython oracle (will be matched by Databricks orchestrator):")
    for k, v in oracle.items():
        print(f"  {k:18s} {v:>20,.2f}")


if __name__ == "__main__":
    main()
